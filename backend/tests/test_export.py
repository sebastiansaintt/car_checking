import io
from openpyxl import load_workbook
import httpx

def test_export_flow():
    client_gerente = httpx.Client(base_url="http://127.0.0.1:8000")

    print("--- 1. Login como Gerente ---")
    login_res = client_gerente.post("/api/auth/login", json={
        "email": "gerente@carchecking.com",
        "password": "gerente123"
    })
    assert login_res.status_code == 200
    print("Login exitoso de Gerente.")

    print("\n--- 2. Exportar Inspecciones a Excel ---")
    export_res = client_gerente.get("/api/export/inspecciones")
    assert export_res.status_code == 200, f"Falló exportación: {export_res.text}"
    assert export_res.headers.get("content-type") == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment; filename=reporte_inspecciones_" in export_res.headers.get("content-disposition", "")
    print(f"Header Content-Disposition: {export_res.headers.get('content-disposition')}")

    # 3. Abrir buffer y validar estructura con openpyxl
    file_bytes = io.BytesIO(export_res.content)
    wb = load_workbook(file_bytes)
    sheet_names = wb.sheetnames
    print(f"Hojas en el Excel generado: {sheet_names}")
    assert "Inspecciones" in sheet_names
    assert "Detalle Checklist" in sheet_names

    ws_resumen = wb["Inspecciones"]
    headers = [cell.value for cell in ws_resumen[1]]
    print(f"Encabezados de Hoja 1: {headers}")
    assert "ID Inspección" in headers
    assert "Patente" in headers
    assert "Resultado General" in headers

    ws_checklist = wb["Detalle Checklist"]
    headers_chk = [cell.value for cell in ws_checklist[1]]
    print(f"Encabezados de Hoja 2: {headers_chk}")
    assert "Item Evaluado" in headers_chk

    print("\n¡Prueba de Exportación a Excel y estructura .xlsx validada con éxito al 100%!")

if __name__ == "__main__":
    test_export_flow()
