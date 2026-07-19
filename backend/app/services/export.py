import io
import uuid
from datetime import datetime
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session

from app.models.usuario import Usuario
from app.repositories.inspeccion import InspeccionRepository
from app.repositories.audit_log import AuditLogRepository

class ExportService:
    @staticmethod
    def generate_inspecciones_excel(
        db: Session,
        usuario: Usuario,
        ip: str,
        vehiculo_id: Optional[uuid.UUID] = None,
        coordinador_id: Optional[uuid.UUID] = None,
        resultado_general: Optional[str] = None,
        fecha_inicio: Optional[datetime] = None,
        fecha_fin: Optional[datetime] = None
    ) -> io.BytesIO:
        """
        Obtiene las inspecciones filtradas, construye un libro Excel en memoria (.xlsx)
        usando openpyxl y registra la exportación en el AuditLog.
        """
        # 1. Consultar inspecciones activas aplicando los filtros
        inspecciones = InspeccionRepository.get_all_active(
            db=db,
            skip=0,
            limit=1000,
            vehiculo_id=vehiculo_id,
            coordinador_id=coordinador_id,
            resultado_general=resultado_general,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin
        )

        # 2. Crear libro y hojas
        wb = Workbook()
        
        # Estilos generales
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # --- Hoja 1: Resumen de Inspecciones ---
        ws_resumen = wb.active
        ws_resumen.title = "Inspecciones"

        headers_resumen = [
            "ID Inspección",
            "Fecha",
            "Patente",
            "Marca",
            "Modelo",
            "Año",
            "Coordinador",
            "Kilometraje (Km)",
            "Resultado General",
            "Mantenimiento Recomendado",
            "Observaciones"
        ]

        ws_resumen.append(headers_resumen)
        
        # Aplicar formato a encabezados
        for col in range(1, len(headers_resumen) + 1):
            cell = ws_resumen.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Insertar datos
        for row_idx, ins in enumerate(inspecciones, start=2):
            veh = ins.vehiculo
            coord = ins.coordinador
            row_data = [
                str(ins.id),
                ins.fecha.strftime("%Y-%m-%d %H:%M"),
                veh.patente if veh else "N/A",
                veh.marca if veh else "N/A",
                veh.modelo if veh else "N/A",
                veh.año if veh else "N/A",
                coord.nombre if coord else "N/A",
                ins.kilometraje,
                ins.resultado_general.upper(),
                ins.mantenimiento_recomendado or "N/A",
                ins.observaciones or "N/A"
            ]
            ws_resumen.append(row_data)

            # Formato de celda por fila
            for col in range(1, len(row_data) + 1):
                c = ws_resumen.cell(row=row_idx, column=col)
                c.border = thin_border
                if col in [1, 2, 3, 6, 8, 9]:
                    c.alignment = center_align
                else:
                    c.alignment = left_align

        # Ajustar ancho de columnas
        for col in ws_resumen.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws_resumen.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # --- Hoja 2: Detalle de Checklist ---
        ws_checklist = wb.create_sheet(title="Detalle Checklist")
        headers_checklist = [
            "ID Inspección",
            "Patente",
            "Item Evaluado",
            "Resultado ("
            "Bueno/Regular/Malo)"
        ]
        ws_checklist.append(["ID Inspección", "Patente", "Item Evaluado", "Resultado"])
        
        for col in range(1, 5):
            cell = ws_checklist.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        chk_row = 2
        for ins in inspecciones:
            veh = ins.vehiculo
            for item in ins.checklist_items:
                catalogo_nombre = item.catalogo.nombre if item.catalogo else "Item"
                ws_checklist.append([
                    str(ins.id),
                    veh.patente if veh else "N/A",
                    catalogo_nombre.capitalize(),
                    item.valor.upper()
                ])
                for col in range(1, 5):
                    c = ws_checklist.cell(row=chk_row, column=col)
                    c.border = thin_border
                    c.alignment = center_align if col != 3 else left_align
                chk_row += 1

        for col in ws_checklist.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws_checklist.column_dimensions[col_letter].width = max(max_len + 3, 15)

        # 3. Guardar libro en buffer de memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 4. Registrar en AuditLog (Requisito explícito de la arquitectura)
        filtros_aplicados = {
            "vehiculo_id": str(vehiculo_id) if vehiculo_id else None,
            "coordinador_id": str(coordinador_id) if coordinador_id else None,
            "resultado_general": resultado_general,
            "fecha_inicio": fecha_inicio.isoformat() if fecha_inicio else None,
            "fecha_fin": fecha_fin.isoformat() if fecha_fin else None,
            "total_registros_exportados": len(inspecciones)
        }

        AuditLogRepository.create_log(
            db=db,
            usuario_id=usuario.id,
            accion="exportar",
            entidad="inspecciones",
            entidad_id=None,
            ip=ip,
            detalle=filtros_aplicados
        )

        return output
